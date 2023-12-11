from customtkinter import *
from tkinter import *
import os
import datetime
from openpyxl import Workbook, load_workbook
from tkcalendar import Calendar
from tkinter import messagebox
from openpyxl.styles import Alignment

lista_parcela = list(range(1, 13))
itens_dict = {}

def data_atual():
    data_hj = datetime.datetime.now()
    pegar_ano = data_hj.year
    pegar_mes = data_hj.month
    pegar_dia = data_hj.day
    return pegar_ano, pegar_mes, pegar_dia



def abrir_calendario():
    def pegar_data():
        global  mes_selecionado, ano_selecionado
        data_selecionada = cal.get_date()
        data_selecionada = datetime.datetime.strptime((data_selecionada), "%m/%d/%y")

        mes_selecionado = data_selecionada.strftime("%m")
        ano_selecionado = data_selecionada.strftime("%Y")
        data_selecionada = data_selecionada.strftime("%d/%m/%y")
        stringvar_data.set(data_selecionada)
        janela_calendario.destroy()


    ano_atual,mes_atual, dia_atual = data_atual()
    janela_calendario = Toplevel(root)
    janela_calendario.title("selecione a data")
    cal = Calendar(janela_calendario, selectmode="day", year=ano_atual, month=mes_atual, day=dia_atual)
    cal.grid(row=0, column=0)
    btn_selecionar = Button(janela_calendario, text="selecionar data", command=pegar_data)
    btn_selecionar.grid(row=1,column=0)


def criar_pasta():
    global ano_selecionado, caminho_da_pasta_ano
    caminho_pasta = r"C:\\Users\\T-GAMER\\Desktop\\gastosplanilhas"
    nome_pasta = ano_selecionado
    caminho_da_pasta_ano = os.path.join(caminho_pasta, nome_pasta)

    try:
        os.mkdir(caminho_da_pasta_ano)
        print(f"Pasta '{nome_pasta}' criada em '{caminho_pasta}'")
    except FileExistsError:
        print(f"A pasta '{nome_pasta}' já existe em '{caminho_pasta}'")

    except Exception as e:
        print(f"Ocorreu um erro: {e}")


def campo_vazio(var):
    var1 = var.get()
    if len(var1) == 0:
        return True
    return False


def validar_numero():
    valor = entry_valor_da_divida.get()
    try:
        valor = valor.replace(",", ".")
        numero = float(valor)
        return numero
    except ValueError:
        messagebox.showerror("nao aceito", "o valor digitado nao é um numero por favor digite corretamente")


def salvar_divida():


    def pegar_dados():
        vazio = []
        global mes_selecionado
        if campo_vazio(entry_nome_da_divida):
            vazio.append("Nome da Divida")
        if campo_vazio(entry_valor_da_divida):
            vazio.append("Valor da Divida")
        if campo_vazio(stringvar_data):
            vazio.append("Data de Pagamento")
        if vazio:
            mensagem = f"campos vazios: {', '.join(vazio)}. preencha todos os campos e tente novamente"
            messagebox.showerror("Campos Vazios", mensagem)

        else:
            quantidade_de_parcela = int(menu_de_opcao.get())
            valor_da_divida = validar_numero()
            nome_da_divida = entry_nome_da_divida.get()
            data_de_pagamento = stringvar_data.get()
            print(nome_da_divida, valor_da_divida, data_de_pagamento, quantidade_de_parcela, mes_selecionado)

            return nome_da_divida, valor_da_divida, data_de_pagamento, quantidade_de_parcela, mes_selecionado

    def criar_ou_abrir_tabela():
        def largura_coluna(tam):
            tamanho = ['A', 'B', 'C', 'D', 'E']
            for i in tamanho:
                ws.column_dimensions[f"{i}"].width = tam


        dados = pegar_dados()
        pasta_de_save = criar_pasta()
        global caminho_da_pasta_ano
        nome_da_divida = dados[0]
        valor_da_divida = dados[1]
        data_de_pagamento = dados[2]
        quantidade_de_parcela = dados[3]

        data_atual_formatada = datetime.datetime.now().strftime("%d-%m-%y")
        caminho_pasta = caminho_da_pasta_ano
        nome_arquivo = f"Gastos_do_mes_{mes_selecionado}.xlsx"
        caminho_completo = os.path.join(caminho_pasta, nome_arquivo)

        columns = ["Data Atual", "Nome da Divida", "Valor da Divida", "Data de Pagamento", "Quantidade de Parcelas"]

        nova_linha = {
            "Data Atual": data_atual_formatada,
            "Nome da Divida": nome_da_divida,
            "Valor da Divida": valor_da_divida,
            "Data de Pagamento": data_de_pagamento,
            "Quantidade de Parcelas": quantidade_de_parcela
        }

        if os.path.exists(caminho_completo):
            wb = load_workbook(caminho_completo)
            ws = wb.active
            largura_coluna(40)
            if not ws['A1'].value:
                ws.append(columns)

            nova_linha_valores = [nova_linha[coluna] for coluna in columns]
            ws.append(nova_linha_valores)
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')

            wb.save(caminho_completo)
        else:
            wb = Workbook()
            ws = wb.active
            largura_coluna(40)
            ws.append(columns)
            nova_linha_values = [nova_linha[coluna] for coluna in columns]
            ws.append(nova_linha_values)
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
            wb.save(caminho_completo)
            print(f'O arquivo {nome_arquivo} foi criado com sucesso')

    criar_ou_abrir_tabela()
    messagebox.showinfo("concluido", " A divida foi salva com sucesso")


def abrir_e_selecionar():
    import os
    from tkinter import filedialog

    def select_dir():
        caminho_da_pasta = r"C:\Users\T-GAMER\Desktop\gastosplanilhas"  # Substitua pelo seu caminho
        diretorio = filedialog.askdirectory(initialdir=caminho_da_pasta)
        return diretorio

    def abrindoplanilha():
        caminho = select_dir()

        if os.path.exists(caminho):
            try:
                if os.path.exists(caminho):
                    os.startfile(caminho)
                else:
                    print("o caminho nao existe")
            except Exception as e:
                print(f"erro ao abrir o arquivo: {e}")

        else:
            print('o arquivo nao existe.')

    abrindoplanilha()


lista_parcela = list(range(1, 13))
lista_parcela_str = [str(i) for i in lista_parcela ]

itens_dict = {}


root = CTk()
root.title("Cadastro de Dívida")
root.geometry("400x500")
root._set_appearance_mode("dark")

stringvar_data = StringVar(root,)
stringvar_data.set("DD-MM-YYYY")

#nome da divida
label_nome_da_divida = CTkLabel(root, text='Nome da Divida')
label_nome_da_divida.pack(padx=10, pady=10)

entry_nome_da_divida = CTkEntry(root, placeholder_text="Ex: peças do carro")
entry_nome_da_divida.pack(padx=10, pady=10)

#valor da divida
label_valor_da_divida = CTkLabel(root, text="Valor da divida")
label_valor_da_divida.pack(padx=10, pady=10)

entry_valor_da_divida = CTkEntry(root, placeholder_text="EX: 1250,00")
entry_valor_da_divida.pack(padx=10, pady=10)

#data de pagamento
button_data_de_pagamento = CTkButton(root, text="Data de Pagamento", command=abrir_calendario)
button_data_de_pagamento.pack(padx=10, pady=10)

label_data_de_pagamento = CTkLabel(root, textvariable=stringvar_data, width=17, )
label_data_de_pagamento.pack(padx=10, pady=10)

#parcelamento
label_parcelado = CTkLabel(root, text="Parcelas")
label_parcelado.pack(padx=10, pady=10)

menu_de_opcao = CTkOptionMenu(root, values=lista_parcela_str)
menu_de_opcao.pack(padx=10, pady=10)
menu_de_opcao.configure(width=15)


#abrir planilhas
button_planilhas = CTkButton(root, text="Abrir Pastas", command=abrir_e_selecionar)
button_planilhas.pack(padx=10, pady=10)

#salvar planilhas

button_salvar_dados = CTkButton(root, text="Salvar Divida", command=salvar_divida)
button_salvar_dados.pack(padx=10, pady=10)

root.mainloop()
